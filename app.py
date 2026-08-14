import os
import io
import uuid
import threading
from datetime import datetime

from functools import wraps
from flask import Flask, request, redirect, url_for, render_template, jsonify, flash, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
from sqlalchemy.pool import NullPool
import qrcode
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-this-secret-key')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///event.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# NullPool = no connection reuse. Without this, an already-open SQLite
# connection can keep reading the OLD file's data after /admin/upload-db
# replaces event.db on disk (the OS lets the old open handle keep pointing
# at the old inode). Opening fresh every query guarantees we always see
# whatever is actually on disk right now.
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'poolclass': NullPool}

# ---------------- Mail configuration (set these in .env) ----------------
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'True') == 'True'
app.config['MAIL_USE_SSL'] = os.environ.get('MAIL_USE_SSL', 'False') == 'True'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', os.environ.get('MAIL_USERNAME'))

# ---------------- Event details (edit or set via .env) ----------------
EVENT_NAME = os.environ.get('EVENT_NAME', 'Piping harmony 2026')
EVENT_DATE = os.environ.get('EVENT_DATE', 'September 05, 2026')
EVENT_VENUE = os.environ.get('EVENT_VENUE', 'Intercontinental,Chennai Mahabalipuram Road, Chennai, Tamil Nadu 600100')
# BASE_URL must be reachable by whatever device/app the organizers use to scan.
# For local testing http://127.0.0.1:5000 is fine. For a real event, deploy the
# app (e.g. on a small VPS or ngrok tunnel) and put that public URL here.
BASE_URL = os.environ.get('BASE_URL', 'http://127.0.0.1:5000')

# Password organizers use to view the participant data (/admin). Change this
# in .env — do not leave the default for a real event.
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')

TEAMS = ['A', 'B', 'C', 'D']

db = SQLAlchemy(app)
mail = Mail(app)

# Simple in-process lock so two simultaneous scans can't both grab the same
# "smallest" team. Fine for a single-process dev server. See README for notes
# on scaling this to multiple workers / a real production deployment.
assign_lock = threading.Lock()


class Participant(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    emp_id = db.Column(db.String(30), nullable=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), nullable=False, unique=True)
    phone = db.Column(db.String(20))
    unique_code = db.Column(db.String(36), unique=True, nullable=False)
    team = db.Column(db.String(1), nullable=True)  # 'A' / 'B' / 'C' / 'D' / None
    registered_at = db.Column(db.DateTime, default=datetime.utcnow)
    scanned_at = db.Column(db.DateTime, nullable=True)
    email_sent = db.Column(db.Boolean, default=False)
    question1 = db.Column(db.String(50), nullable=True)  # e.g. T-shirt size
    question2 = db.Column(db.String(50), nullable=True)  # e.g. Food preference
    question3 = db.Column(db.String(50), nullable=True)  # e.g. Session interest

    def to_dict(self):
        return {
            'id': self.id,
            'emp_id': self.emp_id,
            'name': self.name,
            'email': self.email,
            'unique_code': self.unique_code,
            'team': self.team,
            'registered_at': self.registered_at.isoformat() if self.registered_at else None,
            'scanned_at': self.scanned_at.isoformat() if self.scanned_at else None,
        }


with app.app_context():
    db.create_all()
    # Lightweight migration: if an existing event.db was created before the
    # emp_id column existed, add it in place instead of requiring a wipe.
    with db.engine.connect() as conn:
        existing_cols = [row[1] for row in conn.exec_driver_sql("PRAGMA table_info(participant)")]
        if 'emp_id' not in existing_cols:
            conn.exec_driver_sql("ALTER TABLE participant ADD COLUMN emp_id VARCHAR(30)")
        for qcol in ('question1', 'question2', 'question3'):
            if qcol not in existing_cols:
                conn.exec_driver_sql(f"ALTER TABLE participant ADD COLUMN {qcol} VARCHAR(50)")


def safe_filename_part(text: str) -> str:
    """Turns a participant's name/emp_id into a filesystem-safe chunk for
    use in a downloaded filename — strips anything that isn't a letter,
    digit, space, hyphen, or underscore, then swaps spaces for underscores."""
    cleaned = ''.join(c for c in text if c.isalnum() or c in (' ', '-', '_')).strip()
    return cleaned.replace(' ', '_') or 'participant'


def generate_qr_image(data: str, caption_top: str = None, caption_bottom: str = None) -> io.BytesIO:
    """Generates a QR code. If caption_top/caption_bottom are given, the QR
    is placed on a larger white card with that text printed above/below it
    (e.g. event name on top, participant name on bottom) instead of being a
    bare black-and-white square."""
    from PIL import Image, ImageDraw, ImageFont

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    if not caption_top and not caption_bottom:
        buf = io.BytesIO()
        qr_img.save(buf, format='PNG')
        buf.seek(0)
        return buf

    qr_w, qr_h = qr_img.size
    pad = 40
    top_h = 60 if caption_top else 0
    bottom_h = 50 if caption_bottom else 0
    card_w = qr_w + pad * 2
    card_h = qr_h + pad * 2 + top_h + bottom_h

    card = Image.new("RGB", (card_w, card_h), "white")
    draw = ImageDraw.Draw(card)

    try:
        font_top = ImageFont.truetype("DejaVuSans-Bold.ttf", 26)
        font_bottom = ImageFont.truetype("DejaVuSans.ttf", 20)
    except Exception:
        font_top = ImageFont.load_default()
        font_bottom = ImageFont.load_default()

    y = pad
    if caption_top:
        bbox = draw.textbbox((0, 0), caption_top, font=font_top)
        text_w = bbox[2] - bbox[0]
        draw.text(((card_w - text_w) / 2, y), caption_top, fill="#111827", font=font_top)
        y += top_h

    card.paste(qr_img, (pad, y))
    y += qr_h + 10

    if caption_bottom:
        bbox = draw.textbbox((0, 0), caption_bottom, font=font_bottom)
        text_w = bbox[2] - bbox[0]
        draw.text(((card_w - text_w) / 2, y), caption_bottom, fill="#374151", font=font_bottom)

    buf = io.BytesIO()
    card.save(buf, format='PNG')
    buf.seek(0)
    return buf


def send_qr_email(participant: 'Participant') -> bool:
    """Emails the participant a QR code that encodes the /scan/<code> URL,
    so any organizer's phone camera / QR app can open it directly."""
    scan_url = f"{BASE_URL}/scan/{participant.unique_code}"
    qr_buf = generate_qr_image(scan_url, caption_top=EVENT_NAME, caption_bottom=participant.name)

    subject = f"Your QR Code for {EVENT_NAME}"
    body = f"""Hi {participant.name},

Thank you for registering for {EVENT_NAME}!

Employee ID: {participant.emp_id or '-'}

Event Details:
Date: {EVENT_DATE}
Venue: {EVENT_VENUE}

Your unique QR code is attached to this email. Please bring it (on your
phone or printed) on event day. Any organizer, can scan it and you'll be automatically assigned to a team.

Your registration code: {participant.id}

See you there!
"""
    msg = Message(subject=subject, recipients=[participant.email], body=body)
    msg.attach('event_qr_code.png', 'image/png', qr_buf.read())

    try:
        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"Failed to send email to {participant.email}: {e}")
        return False


def pick_smallest_team() -> str:
    """Team with the fewest members. Ties broken alphabetically (A, B, C, D)."""
    counts = {t: Participant.query.filter_by(team=t).count() for t in TEAMS}
    return min(TEAMS, key=lambda t: (counts[t], t))


def process_scan(code: str):
    """Core scan logic shared by the browser route (/scan/<code>) and the
    JSON API used by the webcam scan station (/api/scan/<code>).
    Returns (participant_or_None, newly_assigned: bool)."""
    participant = Participant.query.filter_by(unique_code=code).first()
    if not participant:
        return None, False

    newly_assigned = False
    with assign_lock:
        db.session.refresh(participant)  # guard against races between requests
        if participant.team is None:
            participant.team = pick_smallest_team()
            participant.scanned_at = datetime.utcnow()
            newly_assigned = True
            db.session.commit()

    return participant, newly_assigned


def admin_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get('is_admin'):
            return redirect(url_for('admin_login', next=request.path))
        return view_func(*args, **kwargs)
    return wrapped


# ---------------------------- Routes ----------------------------

@app.route('/')
def index():
    return render_template(
        'register.html',
        event_name=EVENT_NAME,
        event_date=EVENT_DATE,
        event_venue=EVENT_VENUE,
    )


@app.route('/register', methods=['POST'])
def register():
    emp_id = request.form.get('emp_id', '').strip()
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip().lower()
    phone = request.form.get('phone', '').strip()
    question1 = request.form.get('question1', '').strip()
    question2 = request.form.get('question2', '').strip()
    question3 = request.form.get('question3', '').strip()

    if not emp_id or not name or not email:
        flash('Employee ID, name, and email are all required.', 'error')
        return redirect(url_for('index'))

    if Participant.query.filter_by(email=email).first():
        flash('This email is already registered. Check your inbox for your QR code.', 'error')
        return redirect(url_for('index'))

    participant = Participant(
        emp_id=emp_id,
        name=name,
        email=email,
        phone=phone,
        unique_code=str(uuid.uuid4()),
        question1=question1 or None,
        question2=question2 or None,
        question3=question3 or None,
    )
    db.session.add(participant)
    db.session.commit()

    sent = send_qr_email(participant)
    participant.email_sent = sent
    db.session.commit()

    if sent:
        flash(f'QR code sent to {email}. Check your inbox!', 'success')
    else:
        flash('Registered, but the confirmation email could not be sent. '
              'Please contact the organizers.', 'error')

    return redirect(url_for('index'))


@app.route('/scan/<code>')
def scan(code):
    """Organizer-facing endpoint. Opened automatically when a QR code is
    scanned with a phone camera or any QR scanner app (since the QR encodes
    this URL directly). Assigns a team on first scan, just displays it on
    subsequent scans."""
    participant, newly_assigned = process_scan(code)
    if not participant:
        return render_template('scan_result.html', found=False), 404

    return render_template(
        'scan_result.html',
        found=True,
        participant=participant,
        newly_assigned=newly_assigned,
    )


@app.route('/api/scan/<code>')
def api_scan(code):
    """JSON version of the scan endpoint, used by the webcam-based
    /scan-station page so it can scan continuously without navigating away."""
    participant, newly_assigned = process_scan(code)
    if not participant:
        return jsonify({'found': False}), 404
    return jsonify({
        'found': True,
        'name': participant.name,
        'email': participant.email,
        'team': participant.team,
        'newly_assigned': newly_assigned,
    })


@app.route('/scan-station')
def scan_station():
    """A webcam-driven scanning page for a fixed organizer desk (laptop or
    tablet). Continuously scans QR codes via the device camera in-browser,
    without needing a separate scanner app."""
    return render_template('scan_station.html', event_name=EVENT_NAME)


@app.route('/counts')
def counts():
    result = {t: Participant.query.filter_by(team=t).count() for t in TEAMS}
    result['total_registered'] = Participant.query.count()
    result['total_scanned'] = Participant.query.filter(Participant.team.isnot(None)).count()
    result['unscanned'] = result['total_registered'] - result['total_scanned']
    return jsonify(result)


@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html', event_name=EVENT_NAME)


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        password = request.form.get('password', '')
        if password == ADMIN_PASSWORD:
            session['is_admin'] = True
            next_url = request.args.get('next') or url_for('admin')
            return redirect(next_url)
        flash('Incorrect password.', 'error')
    return render_template('admin_login.html', event_name=EVENT_NAME)


@app.route('/admin/logout')
def admin_logout():
    session.pop('is_admin', None)
    return redirect(url_for('admin_login'))


@app.route('/admin')
@admin_required
def admin():
    """Full participant list for organizers: names, emails, assigned team,
    registration/scan timestamps, and whether the confirmation email sent."""
    q = request.args.get('q', '').strip()
    team_filter = request.args.get('team', '').strip().upper()

    query = Participant.query
    if q:
        like = f"%{q}%"
        query = query.filter(db.or_(Participant.name.ilike(like), Participant.email.ilike(like)))
    if team_filter in TEAMS:
        query = query.filter_by(team=team_filter)
    elif team_filter == 'UNASSIGNED':
        query = query.filter(Participant.team.is_(None))
    elif team_filter == 'SCANNED':
        query = query.filter(Participant.scanned_at.isnot(None))

    participants = query.order_by(Participant.registered_at.desc()).all()
    team_counts = {t: Participant.query.filter_by(team=t).count() for t in TEAMS}

    return render_template(
        'admin.html',
        event_name=EVENT_NAME,
        participants=participants,
        team_counts=team_counts,
        q=q,
        team_filter=team_filter,
        total=Participant.query.count(),
    )


@app.route('/admin/upload-db', methods=['POST'])
@admin_required
def admin_upload_db():
    """Lets an organizer restore the database from a previously downloaded
    backup, straight from the browser. The current live database is backed
    up first (never silently overwritten without a safety copy)."""
    import sqlite3
    import tempfile

    uploaded = request.files.get('db_file')
    if not uploaded or uploaded.filename == '':
        flash('No file selected.', 'error')
        return redirect(url_for('admin'))

    # IMPORTANT: don't reconstruct this path manually — Flask-SQLAlchemy
    # resolves a relative 'sqlite:///event.db' URI against app.instance_path
    # (an "instance/" subfolder), not the project root. Asking the engine
    # directly guarantees we always touch the exact same file the live app
    # is actually reading and writing.
    db_path = db.engine.url.database

    # Save the upload to a temp file first and verify it's a real, valid
    # SQLite database with a participant table before touching anything live.
    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.db')
    os.close(tmp_fd)
    uploaded.save(tmp_path)

    try:
        test_conn = sqlite3.connect(tmp_path)
        count = test_conn.execute("SELECT COUNT(*) FROM participant").fetchone()[0]
        test_conn.close()
    except Exception as e:
        os.remove(tmp_path)
        flash(f'Upload rejected — not a valid database file ({e}).', 'error')
        return redirect(url_for('admin'))

    # Safety copy of what's currently live, timestamped, before overwriting.
    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    safety_path = f"{db_path}.before_upload_{timestamp}.bak"
    try:
        if os.path.exists(db_path):
            src = sqlite3.connect(db_path)
            dst = sqlite3.connect(safety_path)
            src.backup(dst)
            src.close()
            dst.close()
    except Exception as e:
        os.remove(tmp_path)
        flash(f'Could not back up current database, aborted for safety ({e}).', 'error')
        return redirect(url_for('admin'))

    # All existing SQLAlchemy connections need to release the file first.
    db.session.remove()
    db.engine.dispose()

    try:
        os.replace(tmp_path, db_path)
    except Exception as e:
        flash(f'Upload saved but could not replace the live database ({e}). '
              f'Check file permissions on {db_path}.', 'error')
        return redirect(url_for('admin'))

    flash(f'Database restored successfully — {count} participant record(s) loaded.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/edit/<int:participant_id>', methods=['GET', 'POST'])
@admin_required
def admin_edit(participant_id):
    """Lets an organizer correct a participant's details (typos in name/
    email/phone/emp_id) or manually change their assigned team."""
    participant = Participant.query.get_or_404(participant_id)

    if request.method == 'POST':
        emp_id = request.form.get('emp_id', '').strip()
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip().lower()
        phone = request.form.get('phone', '').strip()
        question1 = request.form.get('question1', '').strip()
        question2 = request.form.get('question2', '').strip()
        question3 = request.form.get('question3', '').strip()
        team = request.form.get('team', '').strip().upper()
        team = team if team in TEAMS else None

        if not emp_id or not name or not email or not question1 or not question2 or not question3:
            flash('All fields are required except phone and team.', 'error')
            return redirect(url_for('admin_edit', participant_id=participant.id))

        existing = Participant.query.filter(
            Participant.email == email, Participant.id != participant.id
        ).first()
        if existing:
            flash('Another participant already uses that email.', 'error')
            return redirect(url_for('admin_edit', participant_id=participant.id))

        team_changed = participant.team != team

        participant.emp_id = emp_id
        participant.name = name
        participant.email = email
        participant.phone = phone
        participant.question1 = question1
        participant.question2 = question2
        participant.question3 = question3
        participant.team = team

        # Keep scanned_at consistent with whether a team is actually set —
        # so the "Scanned / checked in" filter stays accurate after a
        # manual edit (e.g. clearing a team, or assigning one by hand).
        if team and not participant.scanned_at:
            participant.scanned_at = datetime.utcnow()
        elif not team:
            participant.scanned_at = None

        db.session.commit()
        flash(f'Updated {participant.name}.' + (' Team changed.' if team_changed else ''), 'success')
        return redirect(url_for('admin'))

    return render_template('admin_edit.html', event_name=EVENT_NAME, participant=participant, teams=TEAMS)


@app.route('/admin/qr/<int:participant_id>')
@admin_required
def admin_download_qr(participant_id):
    """Downloads one participant's QR code as a PNG, named after them —
    e.g. for reprinting a lost QR or handing it out at a walk-in desk."""
    participant = Participant.query.get_or_404(participant_id)
    scan_url = f"{BASE_URL}/scan/{participant.unique_code}"
    qr_buf = generate_qr_image(scan_url, caption_top=EVENT_NAME, caption_bottom=participant.name)

    name_part = safe_filename_part(participant.name)
    emp_part = safe_filename_part(participant.emp_id or str(participant.id))
    filename = f"{emp_part}_{name_part}_QR.png"

    return send_file(qr_buf, as_attachment=True, download_name=filename, mimetype='image/png')


@app.route('/admin/qr-all')
@admin_required
def admin_download_all_qr():
    """Downloads every participant's QR code at once as a single ZIP file,
    each PNG named after that participant."""
    import zipfile

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        used_names = set()
        for participant in Participant.query.order_by(Participant.id).all():
            scan_url = f"{BASE_URL}/scan/{participant.unique_code}"
            qr_buf = generate_qr_image(scan_url, caption_top=EVENT_NAME, caption_bottom=participant.name)

            name_part = safe_filename_part(participant.name)
            emp_part = safe_filename_part(participant.emp_id or str(participant.id))
            filename = f"{emp_part}_{name_part}_QR.png"

            # guard against two participants producing the same filename
            final_name = filename
            counter = 2
            while final_name in used_names:
                final_name = f"{emp_part}_{name_part}_QR_{counter}.png"
                counter += 1
            used_names.add(final_name)

            zf.writestr(final_name, qr_buf.read())

    zip_buf.seek(0)
    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    return send_file(
        zip_buf,
        as_attachment=True,
        download_name=f'all_qr_codes_{timestamp}.zip',
        mimetype='application/zip',
    )


@app.route('/admin/export-excel')
@admin_required
def admin_export_excel():
    """Exports all participant data as a formatted .xlsx file — easier to
    share with organizers who want to filter/sort in Excel rather than
    working from the raw .db file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    headers = ['#', 'Emp ID', 'Name', 'Email', 'Phone', 'Team',
               'Transportation', 'Food Preference', 'BEER',
               'Registered At', 'Scanned At', 'Email Sent']
    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    participants = Participant.query.order_by(Participant.id).all()
    for p in participants:
        email = p.email if not p.email.endswith('@placeholder.local') else ''
        ws.append([
            p.id,
            p.emp_id or '',
            p.name,
            email,
            p.phone or '',
            p.team or 'Unassigned',
            p.question1 or '',
            p.question2 or '',
            p.question3 or '',
            p.registered_at.strftime('%d %b %Y, %H:%M') if p.registered_at else '',
            p.scanned_at.strftime('%d %b %Y, %H:%M') if p.scanned_at else '',
            'Sent' if p.email_sent else 'Not sent',
        ])

    widths = [5, 14, 22, 28, 15, 10, 12, 16, 16, 20, 20, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'participants_{timestamp}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/admin/download-db')
@admin_required
def admin_download_db():
    """Lets an organizer download a safe snapshot of the database straight
    from the browser — no SSH key or terminal needed. Uses SQLite's own
    backup API so it's consistent even while the live app is writing to it."""
    import sqlite3
    import tempfile

    # IMPORTANT: don't reconstruct this path manually — Flask-SQLAlchemy
    # resolves a relative 'sqlite:///event.db' URI against app.instance_path
    # (an "instance/" subfolder), not the project root. Asking the engine
    # directly guarantees we always touch the exact same file the live app
    # is actually reading and writing.
    db_path = db.engine.url.database

    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.db')
    os.close(tmp_fd)

    src = sqlite3.connect(db_path)
    dst = sqlite3.connect(tmp_path)
    src.backup(dst)
    src.close()
    dst.close()

    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    return send_file(
        tmp_path,
        as_attachment=True,
        download_name=f'event_backup_{timestamp}.db',
        mimetype='application/x-sqlite3',
    )


@app.route('/admin/add', methods=['POST'])
@admin_required
def admin_add():
    """Lets an organizer manually add a participant from the /admin page —
    e.g. for a walk-in registration or someone who couldn't use the form.
    A QR code is generated and, if email is filled in, sent the same way
    as self-service registration."""
    emp_id = request.form.get('emp_id', '').strip()
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip().lower()
    phone = request.form.get('phone', '').strip()
    question1 = request.form.get('question1', '').strip()
    question2 = request.form.get('question2', '').strip()
    question3 = request.form.get('question3', '').strip()
    team = request.form.get('team', '').strip().upper()
    team = team if team in TEAMS else None

    if not emp_id or not name or not email or not question1 or not question2 or not question3:
        flash('All fields are required except phone and team.', 'error')
        return redirect(url_for('admin'))

    if Participant.query.filter_by(email=email).first():
        flash('A participant with that email already exists.', 'error')
        return redirect(url_for('admin'))

    participant = Participant(
        emp_id=emp_id,
        name=name,
        email=email,
        phone=phone,
        unique_code=str(uuid.uuid4()),
        team=team,
        scanned_at=datetime.utcnow() if team else None,
        question1=question1,
        question2=question2,
        question3=question3,
    )
    db.session.add(participant)
    db.session.commit()

    sent = send_qr_email(participant)
    participant.email_sent = sent
    db.session.commit()

    if sent:
        flash(f'Participant added and QR code sent to {email}.', 'success')
    else:
        flash('Participant added, but the confirmation email could not be sent.', 'error')

    return redirect(url_for('admin'))


@app.route('/admin/resend/<int:participant_id>', methods=['POST'])
@admin_required
def admin_resend(participant_id):
    """Re-sends the QR code email to a participant — for anyone whose first
    email failed, or who just wants it again. Uses their existing unique_code,
    so the QR still points to the same /scan/<code> link and won't create a
    second, different code for the same person."""
    participant = Participant.query.get_or_404(participant_id)

    if not participant.email or participant.email.endswith('@placeholder.local'):
        return redirect(url_for('admin'))

    sent = send_qr_email(participant)
    participant.email_sent = sent
    db.session.commit()

    return redirect(url_for('admin'))


@app.route('/admin/delete/<int:participant_id>', methods=['POST'])
@admin_required
def admin_delete(participant_id):
    """Permanently removes a participant record from the database."""
    participant = Participant.query.get_or_404(participant_id)
    db.session.delete(participant)
    db.session.commit()
    return redirect(url_for('admin'))


if __name__ == '__main__':
    # debug=True is for local development only — turn off in production.
    app.run(debug=True, host='0.0.0.0', port=5000)
